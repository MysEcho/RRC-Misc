import openpyxl
from openpyxl.styles import PatternFill, Font


def read_questions_from_file(file_path):
    questions = []
    start_index = 0

    with open(file_path, "r") as file:
        lines = file.readlines()
        while start_index < len(lines):
            end_index = start_index
            while end_index < len(lines) and lines[end_index].strip() != "":
                end_index += 1

            if end_index > start_index:
                question_type = lines[start_index].strip()
                question = lines[start_index + 1].strip()
                options = []

                if question_type == "MC":
                    correct_option = lines[start_index + 2].strip().split(")")[0].strip()
                    for j in range(start_index + 3, end_index - 1):
                        option = lines[j].strip().split(")")[1].strip()
                        options.append(option)
                elif question_type == "TF":
                    correct_option = lines[start_index + 2].strip().split(")")[0].strip()
                    options.append(lines[start_index + 3].strip().split(")")[1].strip())
                    options.append(lines[start_index + 4].strip().split(")")[1].strip())
                    options.extend(["", ""])
                elif question_type == "MR":
                    correct_option = (
                        lines[start_index + 2].strip().split(")")[0].strip()
                        + ","
                        + lines[start_index + 3].strip().split(")")[0].strip()
                    )
                    for j in range(start_index + 4, end_index - 1):
                        option = lines[j].strip().split(")")[1].strip()
                        options.append(option)
                else:
                    raise Exception(f"Unknown question type '{question_type}' at line {start_index + 1}")

                # Read the reason as the last line of the question block
                reason = lines[end_index - 1].strip()

                questions.append((question_type, question, correct_option, options, reason))

            start_index = end_index + 1

    return questions


def create_excel_quiz(questions):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = [
        "//Question Type",
        "//Points",
        "//Question Text",
        "//Answer Choice 1",
        "//Answer Choice 2",
        "//Answer Choice 3",
        "//Answer Choice 4",
        "//Reason",
    ]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header

    for i, (question_type, question, correct_option, options, reason) in enumerate(questions, start=1):
        sheet.cell(row=i + 1, column=1, value=question_type)
        sheet.cell(row=i + 1, column=2, value=5)  # Constant Points value
        sheet.cell(row=i + 1, column=3, value=question)

        if question_type == "TF":
            correct_index = 0 if correct_option == "a" else 1
        elif question_type == "MR":
            correct_indices = [ord(opt.strip()[0]) - ord("a") for opt in correct_option.split(",")]
        else:
            correct_index = ord(correct_option[0]) - ord("a")

        for j, option in enumerate(options):
            cell = sheet.cell(row=i + 1, column=j + 4)
            if question_type == "MR":
                if j in correct_indices:
                    cell.value = f"* {option}"
                else:
                    cell.value = option
            else:
                if j == correct_index:
                    cell.value = f"* {option}"
                else:
                    cell.value = option
        sheet.cell(row=i + 1, column=8, value=reason)

    for column in sheet.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save("Excel_Sheet.xlsx")
    print("Excel file has been created successfully!")


if __name__ == "__main__":
    questions = read_questions_from_file("questions.txt")
    create_excel_quiz(questions)
